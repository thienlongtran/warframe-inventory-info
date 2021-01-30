from openpyxl import Workbook
import warframe_market_script

def itemListToExcel(item_list):
    wb = Workbook()
    ws = wb.active

    #Fill Each Column
    fillItemNamesColumn(item_list, ws)
    fillItemPlatinumColumn(item_list, ws)
    fillItemDucatColumn(item_list, ws)
    fillItemPlatDucatEfficiencyRatio(item_list, ws)

    wb.save(filename="result.xlsx")

def fillItemNamesColumn(item_list, ws):
    """
    Fills the first column (A) with the names of the items in current inventory.

    Args:
        item_list (list): A list of Warframe items.
        ws (worksheet): Current active worksheet to save results in.

    Returns:
        no value (function saves results directly to active worksheet)
    """
    ws.column_dimensions['A'].width = 35
    ws["A1"] = "Item"
    for i in range(len(item_list)):
        ws["A" + str(i+2)] = item_list[i]

def fillItemPlatinumColumn(item_list, ws):
    """
    Fills the second column (B) with market platinum price of items in column A.

    Args:
        item_list (list): A list of Warframe items.
        ws (worksheet): Current active worksheet to save results in.

    Returns:
        no value (function saves results directly to active worksheet)
    """
    ws["B1"] = "Plat"
    for i in range(len(item_list)):
        ws["B" + str(i+2)] = str(warframe_market_script.getAveragePlatPrice(ws["A"+ str(i+2)].value))

def fillItemDucatColumn(item_list, ws):
    """
    Fills the third column (C) with Ducat value of items in column A.

    Args:
        item_list (list): A list of Warframe items.
        ws (worksheet): Current active worksheet to save results in.

    Returns:
        no value (function saves results directly to active worksheet)
    """
    ws["C1"] = "Ducat"
    for i in range(len(item_list)):
        ws["C" + str(i+2)] = str(warframe_market_script.getDucatPrice(ws["A"+ str(i+2)].value))

def fillItemPlatDucatEfficiencyRatio(item_list, ws):
    """
    Fills the fourth column (D) with the Platinum to Ducat efficiency ratio.

    Args:
        item_list (list): A list of Warframe items.
        ws (worksheet): Current active worksheet to save results in.

    Returns:
        no value (function saves results directly to active worksheet)
    """
    ws.column_dimensions['D'].width = 10
    ws["D1"] = "Plat:Ducat"
    for i in range(len(item_list)):
        ws["D" + str(i+2)] = "1 : " + str(round((float(ws["C" + str(i+2)].value) / float(ws["B" + str(i+2)].value)),2))